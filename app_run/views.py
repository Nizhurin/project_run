from django.conf import settings
from django.db.models import Count, Q
from django.db.models.aggregates import Sum
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.filters import SearchFilter
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet

from app_run.models import Run, User, AthleteInfo, Challenge, Position, CollectibleItem
from app_run.serializers import RunSerializer, UserSerializer, AthleteInfoSerializer, ChallengeSerializer, \
    PositionSerializer, CollectibleItemSerializer

from geopy.distance import geodesic
from openpyxl import load_workbook


def calculate_run_distance(run: Run) -> float:
    points = Position.objects.filter(run=run)

    if points.count() < 2:
        return 0.0

    total_distance = 0.0
    prev_point = points[0]

    for point in points[1:]:
        total_distance += geodesic(
            (float(prev_point.latitude), float(prev_point.longitude)),
            (float(point.latitude), float(point.longitude))
        ).km
        prev_point = point

    return round(total_distance, 3)


@api_view(['GET'])
def company_details(request):
    details = {
        'company_name': settings.COMPANY_NAME,
        'slogan': settings.SLOGAN,
        'contacts': settings.CONTACTS
    }
    return Response(details)

class RunPagination(PageNumberPagination):
    page_size_query_param = 'size'  # Разрешаем изменять количество объектов через query параметр size в url

class UserPagination(PageNumberPagination):
    page_size_query_param = 'size'  # Разрешаем изменять количество объектов через query параметр size в url

class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter] # Указываем какй класс будет использоваться для фильтра
    filterset_fields = ['status', 'athlete'] # Поля, по которым будет происходить фильтрация
    ordering_fields = ['created_at']  # Поля по которым будет возможна сортировка
    pagination_class = RunPagination  # Указываем пагинацию


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all().exclude(is_superuser=True)
    serializer_class = UserSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name'] # Указываем поля по которым будет вестись поиск
    ordering_fields = ['date_joined']  # Поля по которым будет возможна сортировка
    pagination_class = UserPagination  # Указываем пагинацию

    def get_queryset(self):
        qs = self.queryset  # Используем базовый queryset определенный выше, на уровне класса
        user_type = self.request.query_params.get('type', None)  # Получим параметр type
        is_stuff = False
        if user_type:
            if user_type == 'coach':
                is_stuff = True
            qs = qs.filter(is_staff=is_stuff)
        #  подсчет finished-run для каждого пользователя
        qs = qs.annotate(
            runs_finished=Count('run', filter=Q(run__status='finished'))
        )
        return qs

class RunStartAPIView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == 'init':
            run.status = 'in_progress'
            run.save()
            return Response(status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)

class RunStopAPIView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == 'in_progress':
            run.distance = calculate_run_distance(run=run)
            run.status = 'finished'
            run.save()
            # вычисляем статистические данные
            finished_run = Run.objects.filter(
                athlete=run.athlete,
                status='finished'
            ).aggregate(total_finished=Count('id'), total_finish_distance=Sum('distance'))

            if finished_run['total_finished'] == 10:
                # создаем запись в таблице челенджей без дубликатов
                Challenge.objects.get_or_create(
                    athlete=run.athlete,
                    full_name="Сделай 10 Забегов!"
                )

            if finished_run['total_finish_distance'] >= 50:
                Challenge.objects.get_or_create(
                    athlete=run.athlete,
                    full_name="Пробеги 50 километров!"
                )

            return Response(status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)

class AthleteAPIView(APIView):
    def get(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        athlete_info, _ = AthleteInfo.objects.get_or_create(user=user)

        serializer = AthleteInfoSerializer(athlete_info)
        return Response(serializer.data)

    def put(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        athlete_info, _ = AthleteInfo.objects.get_or_create(user=user)

        serializer = AthleteInfoSerializer(
            athlete_info,
            data=request.data,
            partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(serializer.data, status=status.HTTP_201_CREATED)

class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["athlete"]

class PositionViewSet(ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        run_id = self.request.query_params.get("run")
        if run_id:
            qs = qs.filter(run_id=run_id)
        return qs

class CollectibleItemAPIView(APIView):
    def get(self, request):
        queryset = CollectibleItem.objects.all()
        serializer = CollectibleItemSerializer(queryset, many=True)
        return Response(serializer.data)


@api_view(['POST'])
def upload_collectible_item(request):
    HEADER_MAP = {
        "name": "name",
        "uid": "uid",
        "value": "value",
        "latitude": "latitude",
        "longitude": "longitude",
        "url": "picture",
    }

    def clean_value(key: str, value):
        if value is None:
            return None

        # чистимо строки
        if isinstance(value, str):
            v = value.strip()

            # типова проблема з Excel/CSV-експортом: зайвий ; або лапки
            v = v.strip('";\' ')
            if v.endswith(";"):
                v = v[:-1].strip()

            value = v

        # приводимо типи
        if key in ("latitude", "longitude"):
            try:
                return float(value)
            except (TypeError, ValueError):
                return value  # хай впаде валідація

        if key == "value":
            try:
                return int(value)
            except (TypeError, ValueError):
                return value

        return value

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return Response({"detail": "Файл не передан (key: file)"}, status=status.HTTP_400_BAD_REQUEST)

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    # читаємо заголовки
    raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    raw_headers = [h if h is not None else "" for h in raw_headers]

    # будуємо список полів, у які будемо мапити колонки
    mapped_fields = []
    for h in raw_headers:
        h_norm = str(h).strip().lower()
        mapped_fields.append(HEADER_MAP.get(h_norm))  # може бути None, якщо колонка зайва/невідома

    invalid_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # пропускаємо повністю пусті рядки
        if row is None or all(v is None for v in row):
            continue

        # формуємо data для serializer лише по відомих колонках
        data = {}
        for field_name, cell_value in zip(mapped_fields, row):
            if not field_name:
                continue
            data[field_name] = clean_value(field_name, cell_value)

        serializer = CollectibleItemSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
        else:
            # по ТЗ: повертаємо "сирий" рядок (list of lists)
            invalid_rows.append(list(row))

    wb.close()

    return Response(invalid_rows)